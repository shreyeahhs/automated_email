from fastapi import FastAPI, UploadFile, Form, Query
from fastapi.middleware.cors import CORSMiddleware
import smtplib
from email.message import EmailMessage
from openpyxl import load_workbook
import shutil
import os

app = FastAPI()

# Allow frontend to call this backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],  # Replace with frontend origin
    allow_methods=["*"],
    allow_headers=["*"],
)

# Upload Excel File
@app.post("/api/upload_excel")
async def upload_excel(file: UploadFile):
    with open("emails.xlsx", "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return {"message": "File uploaded successfully"}

# Preview Email Template
@app.get("/api/email_preview")
async def email_preview(
    recipient: str = Query("example@example.com"),
    template: str = Query(None)
):
    if not template:
        return {"preview": "No template provided."}
    try:
        preview = template.format(recipient=recipient)
    except Exception as e:
        preview = f"Error formatting template: {e}"
    return {"preview": preview}

# Send Emails
@app.post("/api/send_emails")
async def send_emails(template: str = Form(...)):
    sender_email = "prathamesh.r.jaiswal@gmail.com"
    sender_password = "eygg gkea iorj dget"  # Use secure method in production

    wb = load_workbook("emails.xlsx")
    sheet = wb["Sheet1"]

    status_report = []

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(sender_email, sender_password)

            for row in range(2, sheet.max_row + 1):
                recipient = sheet.cell(row=row, column=1).value
                if not recipient:
                    continue

                msg = EmailMessage()
                msg["Subject"] = "Demo Email from EduDiagno"
                msg["From"] = sender_email
                msg["To"] = recipient

                content = template.format(recipient=recipient)
                msg.add_alternative(content, subtype="html")

                try:
                    smtp.send_message(msg)
                    sheet.cell(row=row, column=2).value = "Sent"
                    status_report.append(f"✅ Sent to {recipient}")
                except Exception as e:
                    sheet.cell(row=row, column=2).value = "Failed"
                    status_report.append(f"❌ Failed to {recipient}: {e}")

        wb.save("emails.xlsx")
    except Exception as e:
        return {"status": [f"❌ Failed to send emails: {e}"]}

    return {"status": status_report}
