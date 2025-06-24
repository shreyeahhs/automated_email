import smtplib
from email.message import EmailMessage
from openpyxl import load_workbook

# Load email template
with open("template_1.txt", "r") as f:
    email_template = f.read()

# Load emails from Excel file
def get_emails_from_excel(file_path):
    wb = load_workbook(file_path)
    sheet = wb["Sheet1"]
    return wb, sheet

# Email credentials
sender_email = "prathamesh.r.jaiswal@gmail.com"
sender_password = "eygg gkea iorj dget"  # Replace with correct app password

# Load Excel and Sheet
wb, sheet = get_emails_from_excel("emails.xlsx")

# Open SMTP
try:
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(sender_email, sender_password)

        for row in range(2, sheet.max_row + 1):
            recipient = sheet.cell(row=row, column=1).value
            if not recipient:
                continue

            msg = EmailMessage()
            msg["Subject"] = "Demo Email from EduDiagno"
            msg["From"] = sender_email
            msg["To"] = recipient

            # Customize and set email content
            personalized_content = email_template.format(recipient=recipient)
            msg.add_alternative(personalized_content, subtype="html")

            try:
                smtp.send_message(msg)
                print(f"✅ Sent to {recipient}")
                sheet.cell(row=row, column=2).value = "Sent Template_1"
            except Exception as e:
                print(f"❌ Failed to send to {recipient}: {e}")
                sheet.cell(row=row, column=2).value = "Failed"

    # Save updated Excel
    wb.save("emails.xlsx")
    print("✅ Email statuses updated in Excel.")

except Exception as e:
    print("❌ Failed to connect or login:", e)
